from django.shortcuts import render,redirect
from django.contrib.auth.models import User
from django.contrib.auth import authenticate,logout,login
from django.contrib.auth.decorators import login_required
from .models import *
from account.models import *
from django.utils import timezone
from django.core.files.storage import FileSystemStorage
from django.http import JsonResponse
import datetime
import openpyxl
# Create your views here.
def Home(request):
    return  render(request,'index.html')
def FacultyLogin(request):
    error = ""
    if request.method == 'POST':
        u = request.POST['uname']
        p = request.POST['psd']
        user = authenticate(username=u, password=p)
        fid=''
        try:
            if user.is_faculty:
                login(request, user)
                error = "no"
                fid = request.user.faculty
            else:
                error = "yes"

        except:
            error = "yes"

        #print(fid)
        d = {'error': error,'fid':fid}
        return render(request, 'faculty_login.html', d)
    return render(request,'faculty_login.html')

def AdminLogin(request):
    error=""
    if request.method=='POST':
        u=request.POST['uname']
        p=request.POST['pwd']
        user = authenticate(username=u,password=p)
        try:
            if user.is_superuser:
                login(request,user)
                error="no"
            else:
                error="yes"
        except:
            error="yes"
        d={'error':error}
        return render(request,'admin_login.html',d)
    return render(request, 'admin_login.html')

def StudentLogin(request):
    error = ""
    if request.method == 'POST':
        u = request.POST['uname']
        p = request.POST['pwd']
        user = authenticate(username=u, password=p)
        sid=''
        try:
            if user.is_student:
                login(request, user)
                error = "no"
                sid = request.user.student
            else:
                error = "yes"
        except:
            error = "yes"

        d = {'error': error,'sid':sid}
        return render(request, 'student_login.html', d)
    return render(request,'student_login.html')

def Admin(request):
    if request.user.is_superuser:
        return render(request,'admin_dashboard.html')
    return redirect('admin_login')
def FacultyPage(request,pk):
    if request.user.is_faculty:
        c=Courses.objects.all()
        fid=request.user.faculty
        print(fid)
        #print(len(fid.enrolled))
        fid=fid.faculty_id
        f=Faculty.objects.get(faculty_id=fid)
        f=f.enrolled.all()

        p = {'c': c,'f':f,'fid':fid}
        return render(request,'faculty_dashboard.html',p)
    else:
        return redirect('faculty_login')
def StudentProfile(request,pk):
    if request.user.is_student:
        s=Student.objects.get(student_id=pk)
        course=s.enrolled.all()
        return render(request,'student_profile.html',{'course':course,'s':s,'sid':pk})
    return redirect('student_login')

def FacultyProfile(request,pk):
    if request.user.is_faculty:
        f=Faculty.objects.get(faculty_id=pk)
        fe=f.enrolled.all()
        return render(request,'faculty_profile.html',{'f':f,'fe':fe})
    return redirect('faculty_login')
def StudentPage(request,pk):
    if request.user.is_student:
        s = Student.objects.get(student_id=pk)
        course=s.enrolled.all()
        sid=pk
        d = {'s': s,'course':course,'sid':sid}
        return render(request,'student_dashboard.html',d)
    else:
        return redirect('student_login')

def CourseFacultyPage(request,pk,fid):
    if request.user.is_faculty:
        c=Courses.objects.get(course_id=pk)
        n=c.course_atts.all()
        count=n.count()
        ass=c.course_assign.all()
        asscount=ass.count()
        quiz=c.get_quizzes()
        f=Faculty.objects.get(faculty_id=fid)
        f=f.enrolled.all()
        d={'c':c,'fid':fid,'count':count,'n':n ,'ass':ass ,'asscount':asscount,'quiz':quiz,'f':f}
        return render(request,'course_faculty_page.html',d)
    else:
        return redirect('faculty_login')


def CourseStudentPage(request,pk,sid):
    if request.user. is_student:
        c = Courses.objects.get(course_id=pk)
        n = c.course_atts.all()
        print(n)
        co = n.count()
        ass = c.course_assign.all()
        asscount = ass.count()
        quiz = c.get_quizzes()
        s=Student.objects.get(student_id=sid)
        course=s.enrolled.all()
        d={'c':c,'sid':sid,'co':co,'n':n,'ass':ass ,'asscount':asscount,'course':course,'quiz':quiz}
        return render(request,'course_student_page.html',d)
    else:
        return redirect('student_login')
@login_required
def AddCourse(request):
    if not request.user.is_superuser:
        return redirect('admin_login')
    error=""
    if request.method=='POST':
        id=request.POST['cid']
        name=request.POST['cname']
        dept = request.POST['dname']
        try:
            c=Courses.objects.create(course_id=id,course_name=name,dep_id=Department.objects.get(dep_id=dept))
            c.save()
            dep=Department.objects.get(dep_id=dept)
            print(dep.course_nos)
            dep.course_nos += 1
            dep.save()
            error="no"
        except:
            error="yes"
        print(error)
    d={'error':error}
    return render(request,'add_course.html',d)

@login_required
def AddDepartment(request):
    if not request.user.is_superuser:
        return redirect('admin_login')
    error=""
    if request.method=='POST':
        id=request.POST['did']
        name=request.POST['dname']
        date=request.POST['date']
        try:
            Department.objects.create(dep_id=id,dep_name=name,since=date)
            error="no"
        except:
            error="yes"
        print(error)
    d={'error':error}
    return render(request,'add_dept.html',d)

def ViewCourse(request):
    if not request.user.is_superuser:
        return redirect('admin_login')
    else:
        c= Courses.objects.all()
        p={'c':c}
        return render(request,'view_course.html',p)

def ViewDept(request):
    if not request.user.is_superuser:
        return redirect('admin_login')
    else:
        dep= Department.objects.all()
        p={'dep':dep}
        return render(request,'view_department.html',p)

def CourseEnroll(request,cid,fid):
    if not request.user.is_faculty:
        return redirect('faculty_login')
    else:
        f= Faculty.objects.get(faculty_id=fid)
        print("ch")
        print(cid)
        c=Courses.objects.get(course_id=cid)
        print("c "+str(c))
        error=""

        if c not in f.enrolled.all():
                f.enrolled.add(c)
                print(f.enrolled)
                f.courses_enrolled += 1
                f.save()
                error = "no"
        else: error="yes"
        print(f.faculty_name)
        if f.faculty_name not in c.course_faculty:
            if c.course_faculty != "":
                c.course_faculty+=","
            c.course_faculty+=f.faculty_name
            c.save()
            error="no"
        else:
            error="yes"

        d={error:'error'}
        return redirect('faculty',fid)

def enroll(excel_data,c):
    error=""
    for row in excel_data:
        for cell in row:
            if cell not in c.course_students.all():
                c.course_students.add(cell)
                c.part_count += 1
                print(c.course_students.all())
                c.save()
            try:
                s=Student.objects.get(student_id=cell)
                s.enrolled.add(c)
                error="no"
            except:
                error="yes"
    return error



def EnrollStudents(request,id,fid):
    if not request.user.is_faculty:
        return redirect('faculty_login')
    if "POST" == request.method:
        excel_file = request.FILES["excel_file"]

        # you may put validations here to check extension or file size

        wb = openpyxl.load_workbook(excel_file)

        # getting a particular sheet by name out of many sheets
        worksheet = wb["Sheet1"]
        print(worksheet)
        # getting all sheets
        #sheets = wb.sheetnames #(list)
        #print(sheets)
        excel_data = list()
        # iterating over the rows and
        # getting value from each cell in row
        for row in worksheet.iter_rows():
            row_data = list()
            for cell in row:
                row_data.append(str(cell.value))
            excel_data.append(row_data)
        c= Courses.objects.get(course_id=id)

        d= enroll(excel_data,c)
        print("error:" +d)
        d1 = {'excel_data':excel_data,'c': c, 'fid': fid,'d':d,'id':id}
        return render(request, 'enroll_students.html',d1 )
    return render(request, 'enroll_students.html', {'id':id,'fid':fid})

def ShowParticipants(request,cid):
    if not request.user.is_faculty and not request.user.is_student:
        return redirect('')
    c=Courses.objects.get(course_id=cid)
    list=c.course_students.all()
    #print(list)
    name=[]
    for  i in list:
        name.append(Student.objects.get(student_id=i).student_name)
    #print(name)
    merged = [(list[i], name[i]) for i in range(0, len(list))]
    return render(request,'show_participants.html',{'list':list,'c':c,'merged':merged})


def GiveAttendance(request,aid,sid,cid):
    if not request.user.is_student:
        return redirect('student_login')
    if request.method=='POST':
        c=Courses.objects.get(course_id=cid)
        a=Attendance.objects.get(att_id=aid)
        print(a.att_code)
        code = request.POST['acode']
        error=""
        print(code+"  c")
        print(a.active)
        if a.active()==True and a.att_code==code:

            s=AttendanceSubmission.objects.get(id=str(sid)+"_"+str(aid))
            if s.status==False:
                s.stamp=timezone.now()
                s.student=Student.objects.get(student_id=sid)
                s.status=True
                s.save()
                error = "no"
            else:
                error="got already"
            #a.att_students.add(sid)

        elif a.att_code==code or a.active()==False:
            error = "time over"
        elif a.active()==True or code=="":
            error = "code mismatch"
        print("error c" +error)
        d={'error':error,'c':c,'sid':sid}
    return render(request,'course_student_page.html',d)
def PostAttendance(request,cid,fid):
    if not request.user.is_faculty:
        return redirect('faculty_login')
    if request.method == 'POST':
        d = request.POST['dur']
        date=request.POST['date']
        day=request.POST['dah']
        c = Courses.objects.get(course_id=cid)
        import uuid
        print(type(d))
        data = uuid.uuid4().hex[:6].upper()
        print(data)
        a=Attendance.objects.create(att_hour=day,att_date=date,att_duration=d,att_code=data,att_start=timezone.now(),att_end=timezone.now()+datetime.timedelta(minutes=int(d)))
        print(a.att_id)
        c.course_atts.add(a)
        print(c.course_atts)
        print("time: ",end='')
        print(a.att_start,a.att_end)
        print(c.course_atts.all())
        c.save()
        for stud in c.course_students.all():
            name=str(stud)+"_"+str(a.att_id)
            s=AttendanceSubmission.objects.create(id=name,student=Student.objects.get(student_id=stud))
            a.sub.add(name)
            s.save()
        a.save()
        #list=c.course_students.split(',')
        #for i in list:
         #   a.att_students.add(i)
        #a.save()
        send={'c':c,'fid':fid}
        return redirect('course_fac_page',cid,fid)

def ViewAttendanceFaculty(request,cid,fid,aid):
    if not request.user.is_faculty:
        return redirect('faculty_login')

    c = Courses.objects.get(course_id=cid)
    a = Attendance.objects.get(att_id=aid)
    sub=a.sub.all()
    print(sub)
    f = Faculty.objects.get(faculty_id=fid)
    f = f.enrolled.all()
    d={'sub': sub,'cid': cid,'fid': fid,'a': a,'f':f}
    return render(request, 'view_attendance_faculty.html', d)

def PostAssignment(request,cid,fid):
    if not request.user.is_faculty:
        return redirect('faculty_login')
    if request.method == 'POST':
        name= request.POST['name']
        desc=request.POST['desc']
        file=request.FILES.get('file')

        due=request.POST['due']
        c = Courses.objects.get(course_id=cid)
        a=Assignment.objects.create(ass_name=name,ass_desc=desc,ass_posted=timezone.now(),ass_due=due)
        print(a.ass_id)
        print(file)
        for stud in c.course_students.all():
            name=str(stud)+"_"+str(a.ass_id)
            print(name)
            s=Submission.objects.create(id=name,student=stud)
            a.submissions.add(s)
            s.save()
        if file!=None:
            fs=FileSystemStorage(location='vclass/static/assignment')
            print(fs.location)
            print("file stored before")
            file_path=fs.save(file.name,file)
            a.ass_file=file_path
            print("file stored")
        a.save()
        c.course_assign.add(a)
        print(c.course_assign)
        print("time: ")
        print(a.ass_posted,a.ass_due)
        print(c.course_assign.all())
        c.save()

        send={'c':c,'fid':fid}
        return redirect('course_fac_page',cid,fid)

def GradeAssignment(request,sid,cid,fid,aid):
    if not request.user.is_faculty:
        return redirect('faculty_login')
    if request.method=='POST':
        m=request.POST['marks']
        s=Submission.objects.get(id=sid)
        print(m)
        s.marks=int(m)
        s.graded=True
        s.save()
    return redirect('view_assignment_faculty',cid,fid,aid)
def ViewAssignmentStudent(request,cid,sid,aid):
    if not request.user.is_student:
        return redirect('student_login')
    error=""
    a = Assignment.objects.get(ass_id=aid)
    s=Student.objects.get(student_id=sid)
    if request.method=='POST' and a.active() == True:

        f=request.FILES.getlist('files')
        s=Submission.objects.get(id=str(sid)+"_"+str(aid))
        s.date=timezone.now()

        for i in f:
            sf=SubmissionFiles.objects.create(sub_files=i)
            s.files.add(sf)
            sf.save()
        s.save()

        a.submissions.add(s)
        a.save()
        return redirect('view_assignment_student',cid,sid,aid)
    elif request.method=='POST' and a.active()==False:
        error="not active"
    c=Courses.objects.get(course_id=cid)
    course = s.enrolled.all()
    files=""
    try:
        sub=a.submissions.get(student=sid)
        print(sub.files.all())
        files=sub.files.all()
    except:
        sub=None
    print(files)
    for i in files:
        print(i.sub_files)
    print(str(a.active)+"=====files here")
    return render(request, 'view_assignment_student.html', {'course':course,'error':error,'files':files,'a':a,'sub': sub, 'cid': cid, 'sid': sid})
def ViewAssignmentFaculty(request,cid,fid,aid):
    if not request.user.is_faculty:
        return redirect('faculty_login')

    c = Courses.objects.get(course_id=cid)
    a = Assignment.objects.get(ass_id=aid)
    sub=a.submissions.all()
    print(sub)
    f = Faculty.objects.get(faculty_id=fid)
    f = f.enrolled.all()
    d={'sub': sub,'cid': cid,'fid': fid,'f':f,'a': a}

    return render(request,'view_assignment_faculty.html',d)

def ViewQuizFaculty(request,qid,cid,fid):
    if not request.user.is_faculty:
        return redirect('faculty_login')
    q=Quiz.objects.get(quiz_id=qid)
    quiz_questions=q.get_all_questions()
    print(quiz_questions)
    c = Courses.objects.get(course_id=cid)
    quizresult=q.get_results();
    f=Faculty.objects.get(faculty_id=fid)
    f=f.enrolled.all()
    print(quizresult)
    d ={'f':f,'q': q,'quiz_questions':quiz_questions,'quizresult':quizresult,'cid':cid,'fid':fid}
    return render(request,'quiz_faculty_page.html',d)

def QuizPage(request,qid,cid,sid):
    if not request.user.is_student:
        return redirect('student_login')
    q=Quiz.objects.get(quiz_id=qid)
    #quiz_questions = q.get_questions()
    #print(quiz_questions)
    d = {'q': q}
    return render(request,'quiz_page.html',d)

def QuizDatapage(request,qid,cid,sid):
    if not request.user.is_student:
        return redirect('student_login')
    q = Quiz.objects.get(quiz_id=qid)
    quiz_questions = q.get_questions()
    time=q.duration
    flag=True
    if timezone.now()>=q.end_time:
        flag=False
    elif(timezone.now()+datetime.timedelta(minutes=int(time))>q.end_time):
        time=int(((q.end_time-timezone.now()).total_seconds())/60)
    print(quiz_questions)
    data=[]
    ques=[]
    for x in quiz_questions:
        temp=[]
        ques.append(x.question)
        temp.append(x.option1)
        temp.append(x.option2)
        temp.append(x.option3)
        temp.append(x.option4)
        data.append(temp)
    print(time)
    d = {'quiz_questions': quiz_questions, 'sid': sid ,'time':time,'flag':flag}
    return JsonResponse({'data': data,'time':time,'ques':ques,'flag':flag})

def quiz_save_data(request,qid,cid,sid):
    if request.user.is_student and request.is_ajax():
        questions=[]
        data=request.POST
        data_=dict(data.lists())
        data_.pop('csrfmiddlewaretoken')
        s=Student.objects.get(student_id=sid)
        score=0
        qi=Quiz.objects.get(quiz_id=qid)
        multiplier=100/qi.no_of_questions
        for q in data_.keys():
            questions.append(q)
            instance=QuizQuestion.objects.get(question=q)
            input=''.join(data_[q])
            print(instance.correct)
            if instance.correct==input:
                score+=1
        score=score*multiplier
        score = round(score, 2)
        r=QuizResult.objects.get(quiz_id=qi,student_id=s)
        r.score=score
        r.stamp=timezone.now()
        r.status=True
        r.save();
        print(score)
        return JsonResponse({'success':'works'})
        #return redirect('view_quiz_student',qid,cid,sid);

def ViewQuizStudent(request,qid,cid,sid):
    if not request.user.is_student:
        return redirect('student_login')
    q=Quiz.objects.get(quiz_id=qid)
    s=Student.objects.get(student_id=sid)
    qresult=QuizResult.objects.get(student_id=s,quiz_id=q)
    s=Student.objects.get(student_id=sid)
    print(qresult)
    course=s.enrolled.all()
    d ={'q': q,'cid':cid,'sid':sid,'qresult':qresult,'course':course}
    return render(request,'quiz_student_page.html',d)
def updatequiz(request,qid,cid,fid):
    if not request.user.is_faculty:
        return redirect('faculty_login')
    if request.method==  'POST':
        q=Quiz.objects.get(quiz_id=qid)
        q. title=request.POST['title']
        q.start_time = request.POST['start_time']
        q.end_time = request.POST['end_time']
        q.duration=request.POST['duration']
        q.no_of_questions = request.POST['no_of_questions']
        q.save()
        print(q.no_of_questions)
        return redirect('view_quiz_faculty',qid,cid,fid);

def deletequestion(request,qid,cid,fid):
    if not request.user.is_faculty:
        return redirect('faculty_login')
    q = QuizQuestion.objects.get(question_id=qid)
    print(q)
    print("success")
    q.delete()
    return redirect('view_quiz_faculty', qid, cid, fid);


def deletequiz(request, qid,cid,fid):
    if not request.user.is_faculty:
        return redirect('faculty_login')
    q = Quiz.objects.get(quiz_id=qid)
    print(q)
    q.delete()
    return redirect('course_fac_page', cid,fid);
def AddQuestions(request,qid,cid,fid):
    if not request.user.is_faculty:
        return redirect('faculty_login')
    if request.method=='POST':
        question = request.POST.getlist('question[]')
        option1 = request.POST.getlist('option1[]')
        option2 = request.POST.getlist('option2[]')
        option3 = request.POST.getlist('option3[]')
        option4 = request.POST.getlist('option4[]')
        correct = request.POST.getlist('correct[]')
        for i in range(len(question)):
            q=QuizQuestion.objects.create(quiz_id=Quiz.objects.get(quiz_id=qid) ,question=question[i],option1=option1[i],option2=option2[i],option3=option3[i],option4=option4[i])
            if correct[i]=="option1":
                q.correct = option1[i]
            elif correct[i]=="option2":
                q.correct = option2[i]
            elif correct[i]=="option3":
                q.correct = option3[i]
            if correct[i]=="option4":
                q.correct = option4[i]
            q.save()


        return redirect('view_quiz_faculty',qid,cid,fid);
def PostQuiz(request,cid,fid):
    if not request.user.is_faculty:
        return redirect('faculty_login')
    if request.method == 'POST':
        title = request.POST['title']
        start = request.POST['start_time']
        end = request.POST['end_time']
        dur = request.POST['duration']
        n = request.POST['no_of_questions']
        q=Quiz.objects.create(title=title,course_id=Courses.objects.get(course_id=cid) ,start_time=start,end_time=end,duration=dur,no_of_questions=n)
        c=Courses.objects.get(course_id=cid)
        for stud in c.course_students.all():
            r=QuizResult.objects.create(quiz_id=q,student_id=stud)
            r.save()
    return redirect('course_fac_page', cid, fid)

def AdminLogout(request):
    if not request.user.is_superuser:
        return redirect('admin_login')
    logout(request)
    return redirect('admin_login')

def FacultyLogout(request):
    if not request.user.is_faculty:
        return redirect('faculty_login')
    logout(request)
    return redirect('faculty_login')

def StudentLogout(request):
    if not request.user.is_student:
        return redirect('student_login')
    logout(request)
    return redirect('student_login')


